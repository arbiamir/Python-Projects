from tkinter import *
import tkinter as tk
from tkinter.ttk import Combobox
from tkinter import messagebox
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib

root =Tk()
root.title("User Data Entry")
root.geometry("700x400+300+200")
root.resizable(False,False)
root.configure(bg="gold")

file = pathlib.Path("Backened_data.xlsx")
if file.exists():
    pass
else:
    file= Workbook()
    sheet=file.active
    sheet["A1"]="Full Name"
    sheet["B1"]="Phone Number"
    sheet["C1"]="Age"
    sheet["D1"]="Gender"
    sheet["E1"]="Address"
    file.save("Backened_data.xlsx")


def submit():
    name=NameValue.get()
    phone=PhoneValue.get()
    age=AgeValue.get()
    gender=gender_combobox.get()
    address=AddressEntry.get(1.0,END)

    file=openpyxl.load_workbook("Backened_data.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1, value=name)
    sheet.cell(column=2,row=sheet.max_row, value=phone)
    sheet.cell(column=3,row=sheet.max_row, value=age)
    sheet.cell(column=4,row=sheet.max_row, value=gender)
    sheet.cell(column=5,row=sheet.max_row, value=address)
    file.save(r"Backened_data.xlsx")

    messagebox.showinfo("info","Details added!")
    
    NameValue.set("")
    PhoneValue.set("")
    AgeValue.set("")
    AddressEntry.delete(1.0,END)


def clear ():
    NameValue.set("")
    PhoneValue.set("")
    AgeValue.set("")
    AddressEntry.delete(1.0,END)
 
    

icon_image=PhotoImage(file="RBLOGO.png")
root.iconphoto(False,icon_image)

Label(root,text="Please fill out this form:",font="helvatic 14", bg="white", fg="black").place(x=20,y=20)
Label(root,text="Name",font="24",bg="white",fg="black").place(x=50,y=100)
Label(root,text="Phone",font="24",bg="white",fg="black").place(x=50,y=150)
Label(root,text="Age",font="24",bg="white",fg="black").place(x=50,y=200)
Label(root,text="Gneder",font="24",bg="white",fg="black").place(x=370,y=200)
Label(root,text="Address",font="24",bg="white",fg="black").place(x=50,y=250)

NameValue = StringVar()
PhoneValue = StringVar()
AgeValue = StringVar()

NameEntry = Entry(root, textvariable=NameValue, width=45, bd=2, font=20)
PhoneEntry = Entry(root, textvariable=PhoneValue, width=45, bd=2, font=20)
AgeEntry = Entry(root, textvariable=AgeValue, width=15, bd=2, font=20)

gender_combobox= Combobox(root,values=["Male","Female"],font=14, state="r",width=14)
gender_combobox.place(x=440,y=200)
gender_combobox.set(["Male"])

AddressEntry = Text(root,width=50,height=4,bd=4)

NameEntry.place(x=200,y=100)
PhoneEntry.place(x=200,y=150)
AgeEntry.place(x=200,y=200)
AddressEntry.place(x=200,y=250)

Button(root,text="Submit", bg='red',fg='white',width=15,height=2,command=submit).place(x=200,y=350)
Button(root,text="Clear", bg='red',fg='white',width=15,height=2,command=clear).place(x=340,y=350)
Button(root,text="Exit", bg='red',fg='white',width=15,height=2,command=lambda:root.destroy()).place(x=480,y=350)





root.mainloop()